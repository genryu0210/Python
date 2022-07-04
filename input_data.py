# -*- coding: utf-8 -*-
"""
Created on Fri Jul  1 16:54:55 2022

@author: 406429


"""


import openpyxl
import datetime 

t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
date = now.strftime("%Y%m%d")

request = openpyxl.load_workbook("C:\\Users\\406429\\Downloads\\申請書集約.xlsx")
today_request = openpyxl.load_workbook("00【九電工様】各種申請書_Ver2.8_yyyymmdd.xlsm", data_only=True)
#新規の人たち
new_request  = request["新規申請"]
today_new_request = today_request["新規申請"] 
count1 = 6
for j in range(100,500) :
    try :
        a = int(new_request.cell(row=j, column=1).value)
        for i in range(2,26) :
            today_new_request.cell(row=count1, column=i).value = new_request.cell(row=j, column=i).value
        day = today_new_request.cell(row=count1, column=1).value
        new_species = today_new_request.cell(row=count1, column=7).value
        new_request.cell(row=j, column=1).value = str(day) + str(new_species)
        today_new_request.cell(row=count1, column=1).value = str(day) + str(new_species)
        count1+=1
    #else :
    except :
        continue
request_info = today_request["申請者情報"]
if count1 > 6 :
    request_info.cell(row=12, column=3).value = str(count1-6) + "件"


#利用者変更
change_request  = request["利用者変更"]
today_change_request = today_request["利用者変更"] 
count2 = 6
for j in range(1,200) :
    try :
        a = int(change_request.cell(row=j, column=1).value)
        for i in range(2,31) :
            today_change_request.cell(row=count2, column=i).value = change_request.cell(row=j, column=i).value
        day = today_change_request.cell(row=count2, column=1).value
        change_species = today_change_request.cell(row=count2, column=7).value
        change_request.cell(row=j, column=1).value = str(day) + str(change_species)
        today_change_request.cell(row=count2, column=1).value = str(day) + str(change_species)
        count2+=1
    except :
        continue
if count2 > 6 :
    request_info.cell(row=13, column=3).value = str(count2-6) + "件"



#再キッティング
kitting_request  = request["再キッティング"]
today_kitting_request = today_request["再キッティング"] 
count3 = 6
for j in range(100,500) :
    try :
         a = int(kitting_request.cell(row=j, column=1).value)
         for i in range(2,28) :
            today_kitting_request.cell(row=count3, column=i).value = kitting_request.cell(row=j, column=i).value
         day = today_kitting_request.cell(row=count3, column=1).value
         kitting_species = today_kitting_request.cell(row=count3, column=7).value
         kitting_request.cell(row=j, column=1).value = str(day) + str(kitting_species)
         today_kitting_request.cell(row=count3, column=1).value = str(day) + str(kitting_species)
         count3+=1
    except :
        continue
if count3 > 6 :
    request_info.cell(row=14, column=3).value = str(count3-6) + "件"

#故障交換機手配
fixing_request  = request["故障交換機手配"]
today_fixing_request = today_request["故障交換機手配"] 
count4 = 6
for j in range(50,200) :
    try :
         a = int(fixing_request.cell(row=j, column=1).value)
         for i in range(2,30) :
            today_fixing_request.cell(row=count4, column=i).value = fixing_request.cell(row=j, column=i).value
         day = today_fixing_request.cell(row=count4, column=1).value
         fixing_species = today_fixing_request.cell(row=count4, column=7).value
         fixing_request.cell(row=j, column=1).value = str(day) + str(fixing_species)
         today_fixing_request.cell(row=count4, column=1).value = str(day) + str(fixing_species)
         count4+=1
    except :
        continue
if count4 > 6 :
    request_info.cell(row=15, column=3).value = str(count4-6) + "件"


#返却
return_request  = request["返却"]
today_return_request = today_request["返却"] 
count5 = 6
for j in range(100,500) :
    try :
         a = int(return_request.cell(row=j, column=1).value)
         for i in range(2,23) :
            today_return_request.cell(row=count5, column=i).value = return_request.cell(row=j, column=i).value
         day = today_return_request.cell(row=count5, column=1).value
         return_species = today_return_request.cell(row=count5, column=7).value
         return_request.cell(row=j, column=1).value = str(day) + str(return_species)
         today_return_request.cell(row=count5, column=1).value = str(day) + str(return_species)
         count5+=1
    except :
        continue
if count5 > 6 :
    request_info.cell(row=18, column=3).value = str(count5-6) + "件"

today_request.save("xx【九電工様】各種申請書_Ver2.8_" + date + ".xlsx")
request.save("xx申請書集約.xlsx")