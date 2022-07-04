# -*- coding: utf-8 -*-
"""
Created on Mon Jul  4 10:52:06 2022

@author: 406429
"""

import win32com.client
import openpyxl
import datetime

request = openpyxl.load_workbook("C:\\Users\\406429\\Downloads\\申請書集約.xlsx")
today_request = openpyxl.load_workbook("00【九電工様】各種申請書_Ver2.8_yyyymmdd.xlsm", data_only=True)

request_info = today_request["申請者情報"]
t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
date = now.strftime('%Y年%m月%d日')

#date = request_info.cell(row=4,column=4).value

outlook = win32com.client.Dispatch("Outlook.Application")

mail = outlook.CreateItem(0)

mail.to = '406429@kyudenko.co.jp'
mail.cc = '406429@kyudenko.co.jp'
mail.subject = '各種申請_' + str(date) + '分'
mail.bodyFormat = 1
mail.Attachments.Add("C:\\Users\\406429\\Desktop\\新しいフォルダー (2)\\xx【九電工様】各種申請書_Ver2.8_20220704.xlsx")
mail.body = '''MXモバイリング
各種申請受付　ご担当者さま

いつもお世話になっております。九電工の松本です。

本日分の申請です。
お忙しいところ恐縮ですが、ご確認お願いいたします。


'''


mail.display(True)