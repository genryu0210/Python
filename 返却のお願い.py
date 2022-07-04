# -*- coding: utf-8 -*-
"""
Created on Mon Jul  4 11:41:33 2022

@author: 406429
"""
#あとで全部回すためにカウント、その他名前などをインポート
import input_data as Input
import win32com.client
import openpyxl
import datetime



request = openpyxl.load_workbook("C:\\Users\\406429\\Downloads\\申請書集約.xlsx")
today_request = openpyxl.load_workbook("00【九電工様】各種申請書_Ver2.8_yyyymmdd.xlsm", data_only=True)

request_info = today_request["申請者情報"]
t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
date = now.strftime('%Y年%m月%d日')

outlook = win32com.client.Dispatch("Outlook.Application")
mail = outlook.CreateItem(0)

count = [Input.count2-6,Input.count3-6,Input.count4-6,Input.count5-6]
request = [Input.today_change_request, Input.today_kitting_request, Input.today_fixing_request, Input.today_return_request]
for i in range(4) :
    if i == 0 :
        for j in range(0, count[i]) :
            sheet = request[i]
            phone_num = str(sheet.cell(row=j+6, column=12).value)
            phone_front = phone_num[:3]
            phone_last = phone_num[-4:]
            mail.to = str(sheet.cell(row=j+6, column=11).value)
            mail.cc = 'スマートデバイス運用ML <smart-device-operation@kyudenko365.onmicrosoft.com>'
            mail.subject = '【依頼】端末返送のお願い'
            mail.bodyFormat = 1
            mail.body = str(sheet.cell(row=j+6,column=9).value) + '''さま
            
            お疲れさまです。DX推進部　端末申請受付担当：松本です。
            
            本メール受信後5営業日以内に、下記端末を返送いただきますようお願いいたします。
            
            【社員番号】''' + str(sheet.cell(row=j+6, column=17).value) + '''
            【端末所有者名】''' + str(sheet.cell(row=j+6, column=16).value) + '''
            【回線番号】''' + phone_front + '''-XXXX-''' + phone_last + '''	
            【管理番号】''' + sheet.cell(row=j+6, column=13).value +'''
            
            ●返送先
            〒541-0056　大阪市中央区久太郎町四丁目1番3号
            大阪センタービル7階
            TEL：06-6282-3140　
            MXモバイリング株式会社　広域事業第1部
            九電工ヘルプデスク担当宛
            
            ●返送方法
            ①iCloudにサインインしている場合は、サインアウトしたのち返送するようお願いいたします。
            サインアウト方法は以下リンクを参照してください。
            https://support.apple.com/ja-jp/HT208242
            
            ②付箋紙等に以下の内容を記載したものを端末に添付し、緩衝材に梱包した状態で返送をお願いいたします。
            ・所有者社員番号
            ・所有者名
            ・端末電話番号
            ※端末の初期化は必要ありません。（MDMの仕様上できないようになっているため。）
            
            お忙しいところ大変恐縮ですがよろしくお願いいたします。
            '''
            
            
            mail.display(True)
            
    if i == 1 :
        for j in range(0, count[i]) :
            sheet = request[i]
            phone_num = str(sheet.cell(row=j+6, column=12).value)
            phone_front = phone_num[:3]
            phone_last = phone_num[-4:]
            mail.to = str(sheet.cell(row=j+6, column=26).value)
            mail.cc = 'スマートデバイス運用ML <smart-device-operation@kyudenko365.onmicrosoft.com>'
            mail.subject = '【依頼】端末返送のお願い'
            mail.bodyFormat = 1
            mail.body = str(sheet.cell(row=j+6,column=24).value) + '''さま
            
            お疲れさまです。DX推進部　端末申請受付担当：松本です。
            
            本メール受信後5営業日以内に、下記端末を返送いただきますようお願いいたします。
            
            【社員番号】''' + str(sheet.cell(row=j+6, column=9).value) + '''
            【端末所有者名】''' + str(sheet.cell(row=j+6, column=10).value) + '''
            【回線番号】''' + phone_front + '''-XXXX-''' + phone_last + '''	
            【管理番号】''' + sheet.cell(row=j+6, column=13).value +'''
            
            ●返送先
            〒541-0056　大阪市中央区久太郎町四丁目1番3号
            大阪センタービル7階
            TEL：06-6282-3140　
            MXモバイリング株式会社　広域事業第1部
            九電工ヘルプデスク担当宛
            
            ●返送方法
            ①iCloudにサインインしている場合は、サインアウトしたのち返送するようお願いいたします。
            サインアウト方法は以下リンクを参照してください。
            https://support.apple.com/ja-jp/HT208242
            
            ②付箋紙等に以下の内容を記載したものを端末に添付し、緩衝材に梱包した状態で返送をお願いいたします。
            ・所有者社員番号
            ・所有者名
            ・端末電話番号
            ※端末の初期化は必要ありません。（MDMの仕様上できないようになっているため。）
            
            お忙しいところ大変恐縮ですがよろしくお願いいたします。
            '''
            
            
            mail.display(True)
            
    if i == 2 :
        for j in range(0, count[i]) :
            sheet = request[i]
            phone_num = str(sheet.cell(row=j+6, column=12).value)
            phone_front = phone_num[:3]
            phone_last = phone_num[-4:]
            mail.to = str(sheet.cell(row=j+6, column=28).value)
            mail.cc = 'スマートデバイス運用ML <smart-device-operation@kyudenko365.onmicrosoft.com>'
            mail.subject = '【依頼】端末返送のお願い'
            mail.bodyFormat = 1
            mail.body = str(sheet.cell(row=j+6,column=26).value) + '''さま
            
            お疲れさまです。DX推進部　端末申請受付担当：松本です。
            
            本メール受信後5営業日以内に、下記端末を返送いただきますようお願いいたします。
            
            【社員番号】''' + str(sheet.cell(row=j+6, column=19).value) + '''
            【端末所有者名】''' + str(sheet.cell(row=j+6, column=18).value) + '''
            【回線番号】''' + phone_front + '''-XXXX-''' + phone_last + '''	
            【管理番号】''' + sheet.cell(row=j+6, column=13).value +'''
            
            ●返送先
            〒541-0056　大阪市中央区久太郎町四丁目1番3号
            大阪センタービル7階
            TEL：06-6282-3140　
            MXモバイリング株式会社　広域事業第1部
            九電工ヘルプデスク担当宛
            
            ●返送方法
            ①iCloudにサインインしている場合は、サインアウトしたのち返送するようお願いいたします。
            サインアウト方法は以下リンクを参照してください。
            https://support.apple.com/ja-jp/HT208242
            
            ②付箋紙等に以下の内容を記載したものを端末に添付し、緩衝材に梱包した状態で返送をお願いいたします。
            ・所有者社員番号
            ・所有者名
            ・端末電話番号
            ※端末の初期化は必要ありません。（MDMの仕様上できないようになっているため。）
            
            お忙しいところ大変恐縮ですがよろしくお願いいたします。
            '''
            
            
            mail.display(True)
            
    if i == 3 :
        for j in range(0, count[i]) :
            sheet = request[i]
            phone_num = str(sheet.cell(row=j+6, column=13).value)
            phone_front = phone_num[:3]
            phone_last = phone_num[-4:]
            mail.to = str(sheet.cell(row=j+6, column=11).value)
            mail.cc = 'スマートデバイス運用ML <smart-device-operation@kyudenko365.onmicrosoft.com>'
            mail.subject = '【依頼】端末返送のお願い'
            mail.bodyFormat = 1
            mail.body = str(sheet.cell(row=j+6,column=9).value) + '''さま
            
            お疲れさまです。DX推進部　端末申請受付担当：松本です。
            
            本メール受信後5営業日以内に、下記端末を返送いただきますようお願いいたします。
            
            【社員番号】''' + str(sheet.cell(row=j+6, column=18).value) + '''
            【端末所有者名】''' + str(sheet.cell(row=j+6, column=17).value) + '''
            【回線番号】''' + phone_front + '''-XXXX-''' + phone_last + '''	
            【管理番号】''' + sheet.cell(row=j+6, column=14).value +'''
            
            ●返送先
            〒541-0056　大阪市中央区久太郎町四丁目1番3号
            大阪センタービル7階
            TEL：06-6282-3140　
            MXモバイリング株式会社　広域事業第1部
            九電工ヘルプデスク担当宛
            
            ●返送方法
            ①iCloudにサインインしている場合は、サインアウトしたのち返送するようお願いいたします。
            サインアウト方法は以下リンクを参照してください。
            https://support.apple.com/ja-jp/HT208242
            
            ②付箋紙等に以下の内容を記載したものを端末に添付し、緩衝材に梱包した状態で返送をお願いいたします。
            ・所有者社員番号
            ・所有者名
            ・端末電話番号
            ※端末の初期化は必要ありません。（MDMの仕様上できないようになっているため。）
            
            お忙しいところ大変恐縮ですがよろしくお願いいたします。
            '''
            
            
            mail.display(True)