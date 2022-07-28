import openpyxl
import datetime
import win32com.client


class AlreadyExistError(Exception):
    print("申請書は存在しています")
    pass

t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
date = now.strftime("%Y%m%d")

m_d = str(now.strftime("%m/%d"))

# 提出するexcelの日付更新
excel = win32com.client.Dispatch("Excel.Application")

try:
    with open(r"C:\Users\406429\Downloads\申請書集約.xlsx") as f:
        _ = 0
except FileNotFoundError:
    downloads = excel.Workbooks.open(r"C:\Users\406429\KYUDENKO CORPORATION\DX推進部 - 100_Forms申請\申請書集約.xlsx")
    downloads.SaveAs(r"C:\Users\406429\Downloads\申請書集約.xlsx")

wb = excel.Workbooks.open(r"C:\Users\406429\KYUDENKO CORPORATION\DX推進部 - 100_Forms申請\申請書集約.xlsx")
today_format = excel.Workbooks.Open(r"C:\Users\406429\PycharmProjects\python_programming\MXM様申請\00【九電工様】各種申請書_Ver2.8_yyyymmdd.xlsm")
today_format.Save()
request = openpyxl.load_workbook(r"C:\Users\406429\Downloads\申請書集約.xlsx")
today_request = openpyxl.load_workbook("00【九電工様】各種申請書_Ver2.8_yyyymmdd.xlsm", keep_vba=True, data_only=True)
# 新規の人たち
ws = wb.Worksheets("新規申請")
new_request = request["新規申請"]
today_new_request = today_request["新規申請"]
count1 = 6
for j in range(1, new_request.max_row + 1):

    try:

        a = int(new_request.cell(row=j, column=1).value)
        for i in range(2, 26):
            today_new_request.cell(row=count1, column=i).value = new_request.cell(row=j, column=i).value
        day = today_new_request.cell(row=count1, column=1).value
        new_species = today_new_request.cell(row=count1, column=7).value
        new_request.cell(row=j, column=1).value = str(day) + str(new_species)
        today_new_request.cell(row=count1, column=1).value = str(day) + str(new_species)
        ws.Cells(j, 1).Value = str(day) + str(new_species)
        count1 += 1
    except:
        continue
request_info = today_request["申請者情報"]
if count1 > 6:
    request_info.cell(row=12, column=3).value = str(count1 - 6) + "件"

# 利用者変更
ws = wb.Worksheets(2)
change_request = request["利用者変更"]
today_change_request = today_request["利用者変更"]
count2 = 6
for j in range(1, change_request.max_row + 1):
    try:
        a = int(change_request.cell(row=j, column=1).value)
        for i in range(2, 32):
            today_change_request.cell(row=count2, column=i).value = change_request.cell(row=j, column=i).value
        day = today_change_request.cell(row=count2, column=1).value
        change_species = today_change_request.cell(row=count2, column=7).value
        change_request.cell(row=j, column=1).value = str(day) + str(change_species)
        today_change_request.cell(row=count2, column=1).value = str(day) + str(change_species)
        change_request.cell(row=j, column=32).value = str(m_d)
        ws.Cells(j, 1).Value = str(day) + str(change_species)
        count2 += 1
    except:
        continue
if count2 > 6:
    request_info.cell(row=13, column=3).value = str(count2 - 6) + "件"

# 再キッティング
ws = wb.Worksheets(3)
kitting_request = request["再キッティング"]
today_kitting_request = today_request["再キッティング"]
count3 = 6
for j in range(1, kitting_request.max_row + 1):
    try:
        a = int(kitting_request.cell(row=j, column=1).value)
        for i in range(2, 29):
            today_kitting_request.cell(row=count3, column=i).value = kitting_request.cell(row=j, column=i).value
        day = today_kitting_request.cell(row=count3, column=1).value
        kitting_species = today_kitting_request.cell(row=count3, column=7).value
        kitting_request.cell(row=j, column=1).value = str(day) + str(kitting_species)
        today_kitting_request.cell(row=count3, column=1).value = str(day) + str(kitting_species)
        kitting_request.cell(row=j, column=29).value = str(m_d)
        ws.Cells(j, 1).Value = str(day) + str(kitting_species)
        count3 += 1
    except:
        continue
if count3 > 6:
    request_info.cell(row=14, column=3).value = str(count3 - 6) + "件"

# 故障交換機手配
ws = wb.Worksheets(4)
fixing_request = request["故障交換機手配"]
today_fixing_request = today_request["故障交換機手配"]
count4 = 6
for j in range(1, fixing_request.max_row + 1):
    try:
        a = int(fixing_request.cell(row=j, column=1).value)
        for i in range(2, 31):
            today_fixing_request.cell(row=count4, column=i).value = fixing_request.cell(row=j, column=i).value
        day = today_fixing_request.cell(row=count4, column=1).value
        fixing_species = today_fixing_request.cell(row=count4, column=7).value
        fixing_request.cell(row=j, column=1).value = str(day) + str(fixing_species)
        today_fixing_request.cell(row=count4, column=1).value = str(day) + str(fixing_species)
        fixing_request.cell(row=j, column=32).value = str()
        ws.Cells(j, 1).Value = str(day) + str(fixing_species)
        count4 += 1
    except:
        continue
if count4 > 6:
    request_info.cell(row=15, column=3).value = str(count4 - 6) + "件"

# 返却
ws = wb.Worksheets(6)
return_request = request["返却"]
today_return_request = today_request["返却"]
count5 = 6
for j in range(1, return_request.max_row + 1):
    try:
        a = int(return_request.cell(row=j, column=1).value)
        for i in range(2, 24):
            today_return_request.cell(row=count5, column=i).value = return_request.cell(row=j, column=i).value
        day = today_return_request.cell(row=count5, column=1).value
        return_species = today_return_request.cell(row=count5, column=7).value
        return_request.cell(row=j, column=1).value = str(day) + str(return_species)
        today_return_request.cell(row=count5, column=1).value = str(day) + str(return_species)
        return_request.cell(row=j, column=24).value = str(m_d)
        ws.Cells(j, 1).Value = str(day) + str(return_species)
        count5 += 1
    except:
        continue
if count5 > 6:
    request_info.cell(row=18, column=3).value = str(count5 - 6) + "件"

try:
    with open(r"C:\Users\406429\Dropbox (株式会社九電工)\【スマホPJ】展開運用\60_運用\50_申請_通常\01_MX送付分\【九電工様】各種申請書_Ver2.8_" + date + ".xlsm") as f:
        raise AlreadyExistError
except AlreadyExistError:
    pass

today_request.save(
    r"C:\Users\406429\Dropbox (株式会社九電工)\【スマホPJ】展開運用\60_運用\50_申請_通常\01_MX送付分\【九電工様】各種申請書_Ver2.8_" + date + ".xlsm")
request.save("xx申請書集約.xlsx")

