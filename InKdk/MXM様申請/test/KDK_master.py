import openpyxl
import os
import tqdm
import pandas as pd


def get_workbook(workbook_name):
    try:
        if workbook_name == "KDK_master.py":
            pass
        workbook = openpyxl.load_workbook(workbook_name, data_only=True)

        print(f"{workbook_name}を読み込みました。")
        return workbook
    except:
        print(f"{workbook_name}名を確認してください")


def get_graph(i):
    flag = 0
    for j in range(2, 9135):
        if ws.cell(row=i, column=6).value == m_suishin_sheet.cell(row=j, column=2).value:
            ws.cell(row=i, column=9, value=m_suishin_sheet.cell(row=j, column=9).value)
            ws.cell(row=i, column=10, value=m_suishin_sheet.cell(row=j, column=8).value)
            ws.cell(row=i, column=11, value=m_suishin_sheet.cell(row=j, column=11).value)
            ws.cell(row=i, column=12, value=m_suishin_sheet.cell(row=j, column=10).value)
            flag += 1
        if ws.cell(row=i, column=2).value == phonenum_list_sheet.cell(row=j, column=18).value:
            ws.cell(row=i, column=3, value=phonenum_list_sheet.cell(row=j, column=2).value)
            ws.cell(row=i, column=4, value=phonenum_list_sheet.cell(row=j, column=14).value)
            flag += 1
        if ws.cell(row=i, column=3).value == fromsb_sheet.cell(row=j, column=3).value:
            ws.cell(row=i, column=5, value=fromsb_sheet.cell(row=j, column=26).value)
            flag += 1
        if flag > 3:
            break


path_lists = os.listdir(os.path.dirname(__file__))
workbooks = []
print(path_lists)
sheet_name = ["Access取込用に加工", "M_推進者", "Sheet1", "回線番号情報", "Sheet1"]
for i in range(5):
    workbooks.append(get_workbook(path_lists[i]))

fromsb = workbooks[0]
m_suishin = workbooks[1]
mailaddress_list = workbooks[2]
phonenum_list = workbooks[3]
target = workbooks[4]
fromsb_sheet = fromsb["Access取込用に加工"]
m_suishin_sheet = m_suishin["M_推進者"]
mailaddress_list_sheet = mailaddress_list["Sheet1"]
phonenum_list_sheet = phonenum_list["回線番号情報"]
ws = target["Sheet1"]
ws.cell(row=1, column=3, value="回線番号")
ws.cell(row=1, column=4, value="機種名")
ws.cell(row=1, column=5, value="レンタル課金終了日")

for i in tqdm.tqdm(range(1, 8139)):
    ws.cell(row=i, column=1, value=mailaddress_list_sheet.cell(row=i, column=3).value)
    ws.cell(row=i, column=2, value=mailaddress_list_sheet.cell(row=i, column=2).value)
    ws.cell(row=i, column=6, value=mailaddress_list_sheet.cell(row=i, column=22).value)
    ws.cell(row=i, column=7, value=mailaddress_list_sheet.cell(row=i, column=34).value)
    ws.cell(row=i, column=8, value=mailaddress_list_sheet.cell(row=i, column=23).value)
    flag = 0
    get_graph(i)

target.save("created.xlsx")
target.close()
print("おわり")
