# -*- coding: utf-8 -*-
"""
Created on Thu Jul  7 14:51:05 2022

@author: 406429
"""

from input_send_day import Input
import input_send_day as isd
import win32com.client
import pandas as pd
import openpyxl

request = isd.request
return_requesta = request["返却"]
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)

# 受信ボックスのフォルダを取得
folders = inbox.Folders

new_request = folders("返却端末受領のお知らせ").Items
    
df_mail = pd.DataFrame()
i = 0
for message in new_request: 
    df_mail.loc[i, "receivedtime"] = pd.to_datetime(str(message.ReceivedTime)[:-6])
    df_mail.loc[i, "sender"] = str(message.Sender)
    df_mail.loc[i, "subject"] = str(message.Subject)
    df_mail.loc[i, "body"] = str(message.body)
    print(df_mail.loc[i, "subject"])
    num = df_mail.loc[i, "subject"][1:15]
    print(num)
    for j in range(1, Input.return_request.max_row+1):
        if str(return_requesta.cell(row=j, column=1).value) == str(num):
            return_requesta.cell(row=j, column=25).value = str(Input.m_d + "MXM様受領")
    i += 1
    
request.save("xx申請書集約.xlsx")
