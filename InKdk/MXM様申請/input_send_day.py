# -*- coding: utf-8 -*-
"""
Created on Thu Jul  7 11:29:29 2022

@author: 406429
"""
import openpyxl
import input_data as Input
import win32com.client
import pandas as pd

request = openpyxl.load_workbook("xx申請書集約.xlsx")
request = win32com.client.Dispatch("Excel.Application")
new_requesta = request.Worksheets(1)
change_requesta = request.Worksheets(2)
kitting_requesta = request.Worksheets(3)
fixing_requesta = request.Worksheets(4)
return_requesta = request.Worksheets(6)
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)

# 受信ボックスのフォルダを取得
folders = inbox.Folders
# 名前被り良くない
new_request = folders("新規申請").Items

df_mail = pd.DataFrame()
i = 0
for message in new_request:
    df_mail.loc[i, "receivedtime"] = pd.to_datetime(str(message.ReceivedTime)[:-6])
    df_mail.loc[i, "sender"] = str(message.Sender)
    df_mail.loc[i, "subject"] = str(message.Subject)
    df_mail.loc[i, "body"] = str(message.body)
    print(df_mail.loc[i, "subject"])
    num = df_mail.loc[i, "subject"][1:17]
    for j in range(1, Input.new_request.max_row + 1):
        if str(new_requesta.cells(j, 1).value) == num:
            new_requesta.cells(j, 28).value = Input.m_d
    i += 1

new_request = folders("利用者変更").Items

df_mail = pd.DataFrame()
i = 0
for message in new_request:
    df_mail.loc[i, "receivedtime"] = pd.to_datetime(str(message.ReceivedTime)[:-6])
    df_mail.loc[i, "sender"] = str(message.Sender)
    df_mail.loc[i, "subject"] = str(message.Subject)
    df_mail.loc[i, "body"] = str(message.body)
    print(df_mail.loc[i, "subject"])
    num = df_mail.loc[i, "subject"][1:18]
    for j in range(1, change_requesta.max_row + 1):
        if str(Input.change_request.cells(j, 1).value) == num:
            change_requesta.cells(j, 33).value = Input.m_d
    i += 1

new_request = folders("再キッティング").Items

df_mail = pd.DataFrame()
i = 0
for message in new_request:
    df_mail.loc[i, "receivedtime"] = pd.to_datetime(str(message.ReceivedTime)[:-6])
    df_mail.loc[i, "sender"] = str(message.Sender)
    df_mail.loc[i, "subject"] = str(message.Subject)
    df_mail.loc[i, "body"] = str(message.body)
    print(df_mail.loc[i, "subject"])
    num = df_mail.loc[i, "subject"][1:20]
    for j in range(1, Input.kitting_request.max_row + 1):
        if str(kitting_requesta.cells(j, 1).value) == num:
            kitting_requesta.cells(j, 30).value = Input.m_d
    i += 1

new_request = folders("故障交換機手配").Items

df_mail = pd.DataFrame()
i = 0
for message in new_request:
    df_mail.loc[i, "receivedtime"] = pd.to_datetime(str(message.ReceivedTime)[:-6])
    df_mail.loc[i, "sender"] = str(message.Sender)
    df_mail.loc[i, "subject"] = str(message.Subject)
    df_mail.loc[i, "body"] = str(message.body)
    print(df_mail.loc[i, "subject"])
    num = df_mail.loc[i, "subject"][1:20]
    for j in range(1, Input.fixing_request.max_row + 1):
        if str(fixing_requesta.cells(j, 1).value) == num:
            fixing_requesta.cells(j, 31).value = Input.m_d
    i += 1
# %%
# 返却

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)

# 受信ボックスのフォルダを取得
folders = inbox.Folders

new_request = folders("返却端末受領のお知らせ").Items

df_mail = pd.DataFrame()
i = 0
for message in new_request:
    df_mail.loc[i, "receivedtime"] = pd.to_datetime(str(message.ReceivedTime)[:-6])
    df_mail.loc[i, "sender"] = str(message.Sender)
    df_mail.loc[i, "subject"] = str(message.Subject)
    df_mail.loc[i, "body"] = str(message.body)
    print(df_mail.loc[i, "subject"])
    num = df_mail.loc[i, "subject"][1:15]
    print(num)
    for j in range(1, Input.return_request.max_row + 1):
        if str(return_requesta.cells(j, 1).value) == str(num):
            return_requesta.cells(j, 25).value = str(Input.m_d + "MXM様受領")
    i += 1